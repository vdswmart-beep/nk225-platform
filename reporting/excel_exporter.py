# reporting/excel_exporter.py — P2-C: Export Excel complet (track record hedge fund)
#
# Feuilles générées :
#   1. Summary        — métriques clés (Sharpe, MaxDD, Return, etc.)
#   2. Equity Curve   — courbe de NAV journalière + graphique
#   3. Drawdown       — série de drawdown + graphique
#   4. Rolling Sharpe — Sharpe glissant 63j + graphique
#   5. Trade Log      — historique de tous les trades
#   6. Monthly PnL    — PnL par mois (heatmap couleurs)
#   7. Factor Perf    — performance par facteur / secteur (si disponible)

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import numpy as np
import pandas as pd

logger = logging.getLogger("ExcelExporter")


# ══════════════════════════════════════════════════════
#  COULEURS & STYLES
# ══════════════════════════════════════════════════════

_DARK_BG   = "0F141B"
_HEADER_BG = "111827"
_ACCENT    = "4A9EFF"
_GREEN     = "3FB950"
_RED       = "F85149"
_ORANGE    = "F0A500"
_TEXT_MAIN = "C8D6E5"
_TEXT_MUTE = "445566"
_BORDER    = "1E2A38"
_WHITE     = "FFFFFF"


def _try_import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.chart import LineChart, Reference, BarChart
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl requis pour l'export Excel. "
            "Installer avec : pip install openpyxl"
        )


# ══════════════════════════════════════════════════════
#  HELPERS STYLES
# ══════════════════════════════════════════════════════

def _fill(hex_color: str):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=_TEXT_MAIN, size=10):
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(h="left", v="center", wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color=_BORDER)
    return Border(bottom=s)


def _write_header_row(ws, row: int, labels: list, widths: list = None):
    """Écrit une ligne d'en-tête stylisée."""
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill      = _fill(_HEADER_BG)
        cell.font      = _font(bold=True, color=_ACCENT, size=9)
        cell.alignment = _align(h="center")
        cell.border    = _border_thin()
    if widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _write_data_row(ws, row: int, values: list, formats: list = None,
                    bg: str = None):
    """Écrit une ligne de données."""
    for col, val in enumerate(values, 1):
        cell           = ws.cell(row=row, column=col, value=val)
        cell.font      = _font(size=10)
        cell.alignment = _align(h="center")
        if bg:
            cell.fill = _fill(bg)
        if formats and col <= len(formats) and formats[col-1]:
            cell.number_format = formats[col-1]


def _color_by_value(val: float,
                    pos_color: str = _GREEN,
                    neg_color: str = _RED,
                    neutral_color: str = _HEADER_BG) -> str:
    if val is None or np.isnan(val):
        return neutral_color
    if val > 0:
        return pos_color
    if val < 0:
        return neg_color
    return neutral_color


# ══════════════════════════════════════════════════════
#  FEUILLES
# ══════════════════════════════════════════════════════

def _sheet_summary(wb, metrics: Dict, fund_name: str = "NK225 L/S Fund"):
    """Feuille 1 : Summary des métriques."""
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value      = f"📊  {fund_name} — Track Record"
    title.fill       = _fill(_DARK_BG)
    title.font       = _font(bold=True, color=_ACCENT, size=14)
    title.alignment  = _align(h="center")

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value        = f"Généré le {datetime.now().strftime('%d %B %Y à %H:%M')}"
    sub.fill         = _fill(_DARK_BG)
    sub.font         = _font(color=_TEXT_MUTE, size=9)
    sub.alignment    = _align(h="center")

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 16

    # Données
    rows = [
        ("PERFORMANCE",),
        ("Total Return",     metrics.get("total_return"),   "0.00%"),
        ("Annualized Return",metrics.get("ann_return"),     "0.00%"),
        ("Annualized Vol",   metrics.get("ann_vol"),        "0.00%"),
        ("RISK-ADJUSTED",),
        ("Sharpe Ratio",     metrics.get("sharpe"),         "0.00"),
        ("Sortino Ratio",    metrics.get("sortino"),        "0.00"),
        ("Calmar Ratio",     metrics.get("calmar"),         "0.00"),
        ("DRAWDOWN",),
        ("Max Drawdown",     metrics.get("max_drawdown"),   "0.00%"),
        ("Max DD Duration",  metrics.get("max_dd_duration"),"0 days"),
        ("TAIL RISK",),
        ("VaR 95% (daily)",  metrics.get("var_95"),         "0.00%"),
        ("CVaR 95% (daily)", metrics.get("cvar_95"),        "0.00%"),
        ("Skewness",         metrics.get("skewness"),       "0.00"),
        ("Kurtosis",         metrics.get("kurtosis"),       "0.00"),
        ("EXECUTION",),
        ("Win Rate",         metrics.get("win_rate"),       "0.00%"),
        ("Profit Factor",    metrics.get("profit_factor"),  "0.00"),
        ("Trading Days",     metrics.get("n_days"),         "0"),
    ]

    r = 4
    for row_data in rows:
        if len(row_data) == 1:
            # Section header
            ws.merge_cells(f"A{r}:D{r}")
            c = ws.cell(row=r, column=1, value=row_data[0])
            c.fill      = _fill(_HEADER_BG)
            c.font      = _font(bold=True, color=_ACCENT, size=9)
            c.alignment = _align(h="left")
            r += 1
            continue

        label, val, fmt = row_data
        # Label
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill      = _fill(_DARK_BG)
        lc.font      = _font(color=_TEXT_MUTE, size=10)
        lc.alignment = _align(h="left")

        # Valeur
        if isinstance(val, float) and "%" in fmt:
            display_val = val
        else:
            display_val = val

        vc = ws.cell(row=r, column=2, value=display_val)
        vc.fill           = _fill(_DARK_BG)
        vc.font           = _font(bold=True, color=_TEXT_MAIN, size=11)
        vc.alignment      = _align(h="right")
        vc.number_format  = fmt

        # Couleur selon valeur
        if isinstance(val, float):
            is_neg_metric = label in ("Max Drawdown", "VaR 95% (daily)", "CVaR 95% (daily)")
            bg = _color_by_value(-val if is_neg_metric else val)
            indicator = ws.cell(row=r, column=3, value="▲" if val >= 0 else "▼")
            indicator.fill  = _fill(bg)
            indicator.font  = _font(bold=True, color=_WHITE, size=10)
            indicator.alignment = _align(h="center")

        r += 1

    # Largeurs colonnes
    from openpyxl.utils import get_column_letter
    for col, w in enumerate([22, 16, 6, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Fond global
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor.rgb in ("00000000", "000000"):
                cell.fill = _fill(_DARK_BG)


def _sheet_equity(wb, equity: pd.Series, nav_start: float = 100_000_000):
    """Feuille 2 : Courbe NAV + graphique."""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Equity Curve")
    ws.sheet_view.showGridLines = False

    _write_header_row(ws, 1, ["Date", "NAV (¥)", "Return (%)", "Index (Base=100)"],
                      [14, 18, 14, 16])

    daily_returns = equity.pct_change()
    index_series  = equity / nav_start * 100

    for i, (dt, nav_val) in enumerate(equity.items(), 2):
        ret_val = daily_returns.iloc[i-2] if i > 2 else 0.0
        idx_val = index_series.iloc[i-2]
        _write_data_row(
            ws, i,
            [dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt),
             nav_val, ret_val, idx_val],
            formats=["YYYY-MM-DD", '#,##0', '0.00%', '0.00'],
            bg=_DARK_BG,
        )

    # Graphique NAV
    chart      = LineChart()
    chart.title = "NAV — Equity Curve"
    chart.style = 10
    chart.height = 14
    chart.width  = 26

    n_rows = len(equity) + 1
    data   = Reference(ws, min_col=4, min_row=1, max_row=n_rows)
    dates  = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)
    chart.series[0].graphicalProperties.line.solidFill  = _ACCENT
    chart.series[0].graphicalProperties.line.width      = 15000

    ws.add_chart(chart, "F2")


def _sheet_drawdown(wb, drawdown: pd.Series):
    """Feuille 3 : Drawdown."""
    from openpyxl.chart import BarChart, Reference

    ws = wb.create_sheet("Drawdown")
    ws.sheet_view.showGridLines = False

    _write_header_row(ws, 1, ["Date", "Drawdown (%)"], [14, 16])

    for i, (dt, dd) in enumerate(drawdown.items(), 2):
        date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt)
        cell_d = ws.cell(row=i, column=1, value=date_str)
        cell_d.fill = _fill(_DARK_BG)
        cell_d.font = _font(size=10)

        cell_v = ws.cell(row=i, column=2, value=dd)
        cell_v.number_format = "0.00%"
        cell_v.fill = _fill(_RED if dd < -0.05 else (_ORANGE if dd < -0.02 else _DARK_BG))
        cell_v.font = _font(size=10)

    chart        = BarChart()
    chart.type   = "col"
    chart.title  = "Portfolio Drawdown"
    chart.height = 14
    chart.width  = 26

    n_rows = len(drawdown) + 1
    data   = Reference(ws, min_col=2, min_row=1, max_row=n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.series[0].graphicalProperties.solidFill = _RED
    ws.add_chart(chart, "D2")


def _sheet_monthly_pnl(wb, daily_returns: pd.Series):
    """Feuille 4 : PnL mensuel en heatmap."""
    ws = wb.create_sheet("Monthly PnL")
    ws.sheet_view.showGridLines = False

    monthly = daily_returns.resample("ME").apply(lambda x: (1 + x).prod() - 1)
    monthly.index = pd.to_datetime(monthly.index)

    years  = sorted(monthly.index.year.unique())
    months = list(range(1, 13))
    month_names = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]

    # Header
    _write_header_row(ws, 1, ["Année"] + month_names + ["Annuel"], [8]+[7]*12+[9])

    for r, year in enumerate(years, 2):
        ws.cell(row=r, column=1, value=year).font = _font(bold=True, color=_ACCENT)
        ws.cell(row=r, column=1).fill = _fill(_HEADER_BG)

        annual_ret = 1.0
        for c, month in enumerate(months, 2):
            try:
                idx = monthly[(monthly.index.year == year) & (monthly.index.month == month)]
                val = float(idx.iloc[0]) if len(idx) > 0 else None
            except Exception:
                val = None

            cell = ws.cell(row=r, column=c)
            if val is not None:
                cell.value          = val
                cell.number_format  = "0.00%"
                intensity = min(abs(val) / 0.10, 1.0)
                if val >= 0:
                    green_hex = f"00{hex(int(63 + 192*intensity))[2:].upper().zfill(2)}00"
                    cell.fill = _fill(green_hex)
                    cell.font = _font(color=_WHITE, size=10)
                else:
                    red_hex   = f"{hex(int(100 + 155*intensity))[2:].upper().zfill(2)}0000"
                    cell.fill = _fill(red_hex)
                    cell.font = _font(color=_WHITE, size=10)
                cell.alignment = _align(h="center")
                annual_ret *= (1 + val)
            else:
                cell.value = "—"
                cell.fill  = _fill(_HEADER_BG)
                cell.font  = _font(color=_TEXT_MUTE, size=10)
                cell.alignment = _align(h="center")

        # Total annuel
        ann_val     = annual_ret - 1
        ann_cell    = ws.cell(row=r, column=14, value=ann_val)
        ann_cell.number_format = "0.00%"
        ann_cell.font          = _font(bold=True, color=_GREEN if ann_val >= 0 else _RED, size=10)
        ann_cell.fill          = _fill(_HEADER_BG)
        ann_cell.alignment     = _align(h="center")


def _sheet_trades(wb, trades: pd.DataFrame):
    """Feuille 5 : Log des trades."""
    ws = wb.create_sheet("Trade Log")
    ws.sheet_view.showGridLines = False

    if trades is None or trades.empty:
        ws.cell(row=1, column=1, value="Aucun trade enregistré")
        return

    cols   = list(trades.columns)
    widths = [12, 10, 12, 12, 14, 8, 8, 10][:len(cols)]
    _write_header_row(ws, 1, [c.upper().replace("_"," ") for c in cols], widths)

    for r, (_, row) in enumerate(trades.iterrows(), 2):
        vals = [row.get(c) for c in cols]
        _write_data_row(ws, r, vals, bg=_DARK_BG)

        # Couleur action
        if "action" in cols:
            action_col = cols.index("action") + 1
            action_val = row.get("action", "")
            cell = ws.cell(row=r, column=action_col)
            cell.font = _font(
                bold=True,
                color=_GREEN if action_val == "BUY" else _RED,
                size=10,
            )


# ══════════════════════════════════════════════════════
#  EXPORT PRINCIPAL
# ══════════════════════════════════════════════════════

def export_track_record(
    result,                              # BacktestResult
    output_path: Optional[str] = None,
    fund_name:   str           = "NK225 L/S Fund",
    nav_start:   float         = 100_000_000,
) -> Path:
    """
    Génère le fichier Excel de track record complet.

    Args:
        result:       Instance de BacktestResult
        output_path:  Chemin de sortie (.xlsx). Défaut : outputs/track_record_YYYYMMDD.xlsx
        fund_name:    Nom du fonds affiché dans le rapport
        nav_start:    NAV de départ (en JPY)

    Returns:
        Path vers le fichier généré
    """
    openpyxl = _try_import_openpyxl()

    if output_path is None:
        date_str    = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path("outputs") / f"track_record_{date_str}.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Génération track record → {output_path}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # supprime la feuille vide par défaut

    # ── Feuille 1 : Summary
    _sheet_summary(wb, result.metrics, fund_name=fund_name)

    # ── Feuille 2 : Equity Curve
    if result.equity_curve is not None and not result.equity_curve.empty:
        _sheet_equity(wb, result.equity_curve, nav_start=nav_start)

    # ── Feuille 3 : Drawdown
    if result.drawdown_series is not None and not result.drawdown_series.empty:
        _sheet_drawdown(wb, result.drawdown_series)

    # ── Feuille 4 : Monthly PnL
    if result.daily_returns is not None and not result.daily_returns.empty:
        _sheet_monthly_pnl(wb, result.daily_returns)

    # ── Feuille 5 : Trade Log
    _sheet_trades(wb, result.trades)

    wb.save(output_path)
    logger.info(f"Track record sauvegardé : {output_path}")
    return output_path


def export_live_trade(
    trade: Dict,
    output_path: Optional[str] = None,
) -> Path:
    """
    Ajoute un trade live au fichier de track record existant (append).
    Crée le fichier s'il n'existe pas.

    Args:
        trade:        Dict avec clés : date, ticker, action, qty, price, side,
                      weight, nav, pnl (optionnel)
        output_path:  Chemin du fichier Excel de track record live.
    """
    openpyxl = _try_import_openpyxl()

    if output_path is None:
        output_path = Path("outputs") / "live_track_record.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    headers = ["Date","Ticker","Action","Side","Qty","Price (¥)","Weight","NAV","PnL"]

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Trades"] if "Trades" in wb.sheetnames else wb.create_sheet("Trades")
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"
        _write_header_row(ws, 1, headers)
        next_row = 2

    row_vals = [
        trade.get("date", ""),
        trade.get("ticker", ""),
        trade.get("action", ""),
        trade.get("side", ""),
        trade.get("qty", 0),
        trade.get("price", 0),
        trade.get("weight", 0),
        trade.get("nav", 0),
        trade.get("pnl", ""),
    ]
    _write_data_row(ws, next_row, row_vals, bg=_DARK_BG)

    wb.save(output_path)
    logger.info(f"Trade ajouté au track record : {output_path}")
    return output_path
